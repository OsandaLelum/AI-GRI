import pandas as pd
import openpyxl

def IdleTimeIdent(current_cycle,time,next_cycle,idle_times,idle_start,idle_end,cp_no,idle_minutes):


    if current_cycle == "Idle" and idle_start == '' and next_cycle != 'Idle': 
        idle_start = time
        idle_end = time
        idle_minutes = 1
        idle_times.append([cp_no, idle_start, idle_end, idle_minutes])
        idle_start = ''
        idle_end = ''
        idle_minutes = 0

    elif current_cycle == "Idle" and idle_start == '':
       idle_start = time
       idle_minutes += 1
                    
    elif current_cycle  == "Idle":
        idle_minutes += 1
        if next_cycle != 'Idle':
            idle_end = time
            idle_times.append([cp_no, idle_start, idle_end, idle_minutes])
            idle_start = ''
            idle_end = ''
            idle_minutes = 0

    return idle_times,idle_start,idle_end,cp_no,idle_minutes


# Define a function to identify day or night shift based on the given time
def get_shift(time_str):
    time_obj = datetime.strptime(time_str, '%m/%d/%y %I:%M %p')
    hour = time_obj.hour
    if hour >= 6 and hour < 18:
        return 'day'
    else:
        return 'night'
    

def IdletimeReport(filename, sheet_name):

    df = pd.read_excel(filename, sheet_name=sheet_name,  header=None)
    idle_times = []
    idle_start = ''
    idle_end = ''
    cp_no = sheet_name
    idle_minutes = 0
    day_idle_time = 0

    for i, row in df.iterrows():
        if sheet_name[2:4] == '01':
            current_cycle = row[8]
            next_cycle = df.loc[i+1, 8] if i < len(df)-1 else None
            time = row[1]

            idle_times,idle_start,idle_end,cp_no,idle_minutes = IdleTimeIdent(current_cycle,time,next_cycle,idle_times,idle_start,idle_end,cp_no,idle_minutes)
        
        else:
            current_cycle = row[7]
            next_cycle = df.loc[i+1, 7] if i < len(df)-1 else None
            time = row[1]

            idle_times,idle_start,idle_end,cp_no,idle_minutes = IdleTimeIdent(current_cycle,time,next_cycle,idle_times,idle_start,idle_end,cp_no,idle_minutes)


    df_idle_times = pd.DataFrame(idle_times, columns=['CP No', 'Idle Time Start', 'Idle Time End', 'Idle Minutes'])
    
    # Add the total idle time during day shift at the bottom of the sheet
    df_idle_times.loc[len(df_idle_times)] = ['', 'Day Shift Idle Time Total:', '', day_idle_time]

    return df_idle_times

# Call the function to analyze the curing press data
filename = 'C:\\Users\\osandal\\Downloads\\GRI\\data set 9\\combined_dataset91_new_new_new.xlsx'
book = openpyxl.load_workbook(filename)
new_book = openpyxl.Workbook()
new_book.remove(new_book['Sheet'])

for sheet_name in book.sheetnames:
    df = IdletimeReport(filename, sheet_name)
    new_sheet = new_book.create_sheet(sheet_name)
    new_sheet.append(['CP No', 'Idle Time Start', 'Idle Time End', 'Idle Minutes'])
    for row in df.iterrows():
        new_sheet.append([row[1]["CP No"], row[1]["Idle Time Start"], row[1]["Idle Time End"], row[1]["Idle Minutes"]])

# Save the new Excel file
new_book.save('C:\\Users\\osandal\\Downloads\\GRI\\data set 9\\combined_dataset99_new_new_new.xlsx')