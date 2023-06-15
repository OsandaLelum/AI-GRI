import openpyxl

# Load the Excel workbook
wb = openpyxl.load_workbook("H:\\GRI\CP data\\New folder\\CP01(28).xlsx")

# Select the first sheet
ws = wb.active

# Set the values for the temperature range
temp_lower = 200
temp_upper = 204
temp_column = "C"

# Set the values for the pressure range
pressure_lower = 15
pressure_upper = 16
pressure_column = "D"

# Set the values for the special pressure range
special_pressure_lower = 24.5
special_pressure_upper = 25.5
special_pressure_column = "D"

# Variables to store the start and end time of the special time range
start_time = None
end_time = None

# Loop through all the rows in the columns
for row in ws.iter_rows(values_only=True):
    temp_value = row[0]
    pressure_value = row[1]
    time_value = row[2]

    temp_cell = ws[temp_column + str(row[0].row)]
    pressure_cell = ws[pressure_column + str(row[0].row)]

    # Check if the temperature is outside the range
    if temp_value < temp_lower or temp_value > temp_upper:
        # Highlight the cell with a fill
        temp_cell.fill = openpyxl.styles.PatternFill(start_color="red", end_color="red", fill_type="solid")

    # Check if the pressure is outside the range
    if pressure_value < pressure_lower or pressure_value > pressure_upper:
        # Highlight the cell with a fill
        pressure_cell.fill = openpyxl.styles.PatternFill(start_color="red", end_color="red", fill_type="solid")

    # Check if the pressure is within the special pressure range
    if special_pressure_lower <= pressure_value <= special_pressure_upper:
        # If the start time has not been set, set it to the current time
        if start_time is None:
            start_time = time_value

        # Update the end time to the current time
        end_time = time_value
    else:
        # If the end time has been set, highlight the deviations from the special pressure range within the time range
        if start_time is not None and end_time is not None:
            for i in range(ws.columns[0].row, row[0].row):
                if start_time <= ws.cell(row=i, column=3).value <= end_time:
                    pressure_cell = ws[special_pressure_column + str(i)]
                    pressure_cell.fill = openpyxl.styles.PatternFill(start_color="yellow", end_color="yellow", fill_type="solid")
            start_time = None
            end_time = None

# Save the workbook
wb.save("data_highlighted.xlsx")
