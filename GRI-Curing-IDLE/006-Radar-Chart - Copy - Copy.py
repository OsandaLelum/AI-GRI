import openpyxl
import matplotlib.pyplot as plt
import numpy as np

# Load the Excel file
wb = openpyxl.load_workbook(r'C:\Users\osandal\OneDrive - Global Rubber Industries (Pvt) Ltd\Desktop\Github\AI-GRI\GRI-Curing-IDLE\data set 9\combined_dataset.xlsx')

# Select the worksheet
ws = wb.active

# Read the data from the worksheet and select the desired columns (B to F)
data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    data.append(row[2:7])

# Transpose the data
data_transposed = list(map(list, zip(*data)))

# Extract the column names
column_names = [cell.value for cell in ws[1][2:7]]

# Plot the radar chart
plt.figure(figsize=(8, 6))
angles = np.linspace(0, 2 * np.pi, len(column_names), endpoint=False)
ax = plt.subplot(111, polar=True)
for values in data_transposed:
    values += values[:1]  # Close the loop
    ax.plot(angles, values, marker='o')
    ax.fill(angles, values, alpha=0.25)
ax.set_xticks(angles)
ax.set_xticklabels(column_names)
ax.set_title('Radar Chart')
ax.grid(True)

plt.show()
