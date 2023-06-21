import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

def LoadingAndUnloading(sheet_name):

    Loadingandunloadingtime = [4,5,5,5,5,5,5,5,8,8,8,12,12,12,4,4,5,5,8,8,8,12,12,12,12,12,12,12,12,12,12,8,8,8,8,12,12,12,12,12,12,12]
    return Loadingandunloadingtime[int(sheet_name[2:4]) - 1]

def analyze_curing_press(file_path,sheet_name):
    # Load the excel file into a pandas dataframe
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Set the pressure and temperature columns to variables
    pressure_col = df[" Int pressure"]
    temperature_col = df[" Int temp"]

    # Create a new column to store the deviations
    df['Deviation'] = ''
    
    # Initialize variables for the current cycle and its starting row
    current_cycle = None
    cycle_start = None
    drain_vaccuum = 0
    AllowedLnUL = LoadingAndUnloading(sheet_name)
    LnUL = 0

    # Loop through the rows in the dataframe
    for i, row in df.iterrows():
        pressure = row[' Int pressure']
        temperature = row[' Int temp']

        # Check if the pressure and temperature are in the range for the high pressure steam step
        if 13 <= pressure <= 17.5 and 190 <= temperature:
            # Start a new cycle
            current_cycle = 'high_pressure_steam'
            cycle_start = i
            df.at[i, 'Cycle Step'] = "High Pressure Steam"
            LnUL = 0
            drain_vaccuum = 0
            
        elif current_cycle == 'high_pressure_steam' and (pressure > 20 and temperature < 190):
            # End the current cycle and start a new one for the hot water step
            current_cycle = 'hot_water'

        elif current_cycle == None and pressure > 23 and 80 < temperature < 175: current_cycle = 'hot_water'  

        elif (current_cycle == 'hot_water' and (temperature < 68 or pressure < 20)):
            # End the current cycle
            current_cycle = 'cold_water'

        elif (current_cycle == 'cold_water' and pressure < 3): current_cycle = "Drain or Vaccuum"

        elif (current_cycle == 'Drain or Vaccuum' and drain_vaccuum > 5 and pressure < 5):
            current_cycle = "Loading and Unloading" 
            drain_vaccuum = 0

        elif (current_cycle == 'Drain or Vaccuum' and pressure > 5): current_cycle = None

        elif ((current_cycle == "Loading and Unloading" and LnUL == AllowedLnUL) or (current_cycle == "Loading and Unloading" and pressure > 5)):
            current_cycle = None
            LnUL = 0
  
        # Check if there is a current cycle and if the pressure or temperature are outside of their range
        if current_cycle is not None:
            if (current_cycle == 'high_pressure_steam' and (pressure < 14.5 or pressure > 16.5)):
                df.at[i, 'Deviation'] = 'Deviation'

            if (current_cycle == 'high_pressure_steam' and (temperature < 200 or temperature > 204)):
                df.at[i, 'Deviation'] = 'Deviation'
            
            if (current_cycle == 'hot_water' and (pressure < 24 or pressure > 26)):
                # Set the deviation for this row
                df.at[i, 'Deviation'] = 'Deviation'
                
        if current_cycle == 'hot_water': df.at[i, 'Cycle Step'] = "Hot Water"
        if current_cycle == 'cold_water':  df.at[i, 'Cycle Step'] = "Cold Water"
        if current_cycle == 'Drain or Vaccuum':  
            df.at[i, 'Cycle Step'] = "Drain or Vaccuum"
            drain_vaccuum += 1 
        if current_cycle == "Loading and Unloading" and LnUL < AllowedLnUL:
            df.at[i, 'Cycle Step'] = "Loading and Unloading"
            LnUL += 1
        if current_cycle == None and pressure < 5: df.at[i, 'Cycle Step'] = "Idle" 


    # Return the modified dataframe
    return df

# Call the function to analyze the curing press data
#book = openpyxl.load_workbook('C:\\Users\\osandal\\Downloads\\GRI\\data set 9\\combined_dataset_1.xlsx')

#book = openpyxl.load_workbook('C:/Users/osandal/Downloads/GRI/data set 9/combined_dataset_1.xlsx')


#book = openpyxl.load_workbook('C:\\Users\\osandal\\Downloads\\GRI\\data set 9\\combined_dataset_1.xlsx')
#book = openpyxl.load_workbook(r'C:\Users\osandal\Downloads\GRI\data set 9\combined_dataset_1.xlsx')
book = openpyxl.load_workbook('C:\\Users\\osandal\\Downloads\\AI\\AI-GRI-main\\GRI-Curing-temp\\data set 9\\combined_dataset.xlsx')
#df=pd.DataFrame('C:\Users\osandal\Downloads\GRI\data set 9\combined_dataset_1.xlsx')

sheet_names = book.sheetnames

new_book = openpyxl.Workbook()
new_book.remove(new_book['Sheet'])
for sheet_name in sheet_names:
    sheet = book[sheet_name]
    df = analyze_curing_press('C:\\Users\\osandal\\Downloads\\AI\\AI-GRI-main\\GRI-Curing-temp\\data set 9\\combined_dataset.xlsx',sheet_name)
    
    new_sheet = new_book.create_sheet(sheet_name)
    for row in df.iterrows():
        for col in range(len(row[1])):
            new_sheet.cell(row=row[0]+1, column=col+1, value=row[1][col])

#new_book.save('C:\Users\osandal\Downloads\GRI\data set 9\combined_dataset9_new_new_new.xlsx') 
new_book.save('C:\\Users\\osandal\\Downloads\\AI\\AI-GRI-main\\GRI-Curing-temp\\data set 9\\combined_dataset_new.xlsx')